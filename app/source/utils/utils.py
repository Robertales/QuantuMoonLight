from openpyxl import Workbook
import csv
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier


def numberOfColumns(filename):
    """
    This function return the number of columns of a given dataset

    :param filename: name of the dataset
    :return: number of columns of the given dataset
    :rtype: int
    """
    f = open(filename, 'r')
    reader = csv.reader(f, delimiter=',')
    numCols = len(next(reader))
    f.close()
    return numCols  # Read first line and count columns


def numberOfRows(filename):
    """
    This function return the number of raws of a given dataset

    :param filename: name of the dataset
    :return: number of raws of the given dataset
    :rtype: int
    """
    results = pd.read_csv(filename)

    return len(results)


def createFeatureList(numCols: int):
    """
    This function create a List of string such as ["feature1","feature2",...,"featureN"
    where N is given in input

    :param numCols: number of feature to insert in the list
    :return: List of feature
    :rtype: list
    """
    featureList = []
    for x in range(numCols):
        stringa = "feature{}".format(x + 1)
        featureList.append(stringa)
    return featureList


def classifier(number_of_training_instances):
    c = KNeighborsClassifier(n_neighbors=number_of_training_instances, weights='distance')
    return c


def prepareData(databasePath):
    # READ dataset

    data = pd.read_csv(databasePath)

    # preprocessing
    # replace categorical data
    # da modificare
    # dictionaryCat = {"class": {"Iris-setosa": 0, "Iris-versicolor": 1, "Iris-virginica": 2}}
    # data.replace(dictionaryCat, inplace=True)

    # print(data.head())

    # detecting classes
    target = data.values[:, -1]
    classes = np.unique(target)
    number_of_classes = len(classes)
    number_of_features = len(data.values[1]) - 1

    # print(str(number_of_classes) + " " + str(number_of_features))

    # split dei dati
    random_state_value = 3
    x_train, x_test = train_test_split(data.values, test_size=0.25, random_state=random_state_value, stratify=target)

    # split dei dati di training
    # x_train ,x_valid = train_test_split(x_train,test_size=0.10)

    number_of_total_instances = len(x_train)

    return x_train, x_test, number_of_features, number_of_classes, number_of_total_instances


def writeTxt(fileName, list_values):
    """
    This function write in a file .txt the values stored in a list

    :param fileName: name of the output file
    :param list_values: list of values to write
    :return: name of the output file
    :rtype: str
    """
    f = open(fileName, "w+")
    for c in list_values:
        f.write(str(c) + "\n")
    f.close()
    return fileName


def writeXls(fileName, generations, evaluations, bestfits, times):
    """
    This function is used in Prototype Selection to write important info of the Genetic algorithm in a file .xls

    :param fileName: name of the output file
    :param generations:
    :param evaluations:
    :param bestfits:
    :param times:
    :return: name of the output file
    """
    wb = Workbook()
    ws1 = wb.active

    ws1.cell(column=1, row=1, value="Number of Generations")
    ws1.cell(column=2, row=1, value="Number of Evaluations")
    ws1.cell(column=3, row=1, value="Best Fitness Value")
    ws1.cell(column=4, row=1, value="Time in seconds")

    row = 2
    for g in generations:
        ws1.cell(column=1, row=row, value=g)
        row += 1

    row = 2
    for g in evaluations:
        ws1.cell(column=2, row=row, value=g)
        row += 1

    row = 2
    for g in bestfits:
        ws1.cell(column=3, row=row, value=g)
        row += 1

    row = 2
    for g in times:
        ws1.cell(column=4, row=row, value=g)
        row += 1

    wb.save(filename=fileName)
    wb.close()

    return fileName

